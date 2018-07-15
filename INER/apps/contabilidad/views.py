from __future__ import absolute_import
from django.shortcuts import render, redirect
from django.core.urlresolvers import reverse_lazy
from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic.base import TemplateView
from django.views.generic import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import (
    CreateView,
    UpdateView,
    DeleteView
)
from apps.contabilidad.models import Contabilidad
from openpyxl import Workbook

class ContabilidadList(ListView):
    model = Contabilidad
    template_name = 'contabilidad/cont_list.html'

class ContabilidadDetail(DetailView):
    model = Contabilidad
    template_name = 'contabilidad/cont_detail.html'

class ContabilidadCreation(CreateView):
    model = Contabilidad
    success_url = reverse_lazy('conta:list')
    fields = ['clave_cucop', 'partida', 'descripcion', 'unidad_medida']
    template_name = 'contabilidad/cont_form.html'

class ContabilidadUpdate(UpdateView):
    model = Contabilidad
    success_url = reverse_lazy('conta:list')
    fields = ['clave_cucop', 'partida', 'descripcion', 'unidad_medida']
    template_name = 'contabilidad/cont_edit.html'

class ContabilidadDelete(DeleteView):
    model = Contabilidad
    success_url = reverse_lazy('conta:list')
    template_name = 'contabilidad/cont_delete.html'

#Excel
class ReporteCucopExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        conta = Contabilidad.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'CUCOP'
        ws.merge_cells('B1:E1')

        ws['B3'] = 'CLAVE CUCOP'
        ws['C3'] = 'PARTIDA'
        ws['D3'] = 'DESCRIPCION'
        ws['E3'] = 'UNIDAD DE MEDIDA'
        cont=4

        for con in conta:
            ws.cell(row=cont,column=2).value = con.clave_cucop
            ws.cell(row=cont,column=3).value = con.partida
            ws.cell(row=cont,column=4).value = con.descripcion
            ws.cell(row=cont,column=5).value = con.unidad_medida
            cont = cont + 1

        nombre_archivo ="ReporteExcel.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
